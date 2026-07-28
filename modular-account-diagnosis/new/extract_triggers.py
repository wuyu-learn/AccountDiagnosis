#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extract sidebar navigation and card trigger descriptions into an xlsx."""
import re
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = Path(r"C:\Users\EDY\Desktop\code\HTML\Account Diagnosis\modular-account-diagnosis\new")
SECTIONS_DIR = BASE_DIR / "sections"
SIDEBAR_FILE = BASE_DIR / "partials" / "sidebar.html"
OUTPUT_FILE = BASE_DIR / "卡片导航与触发内容.xlsx"


def parse_sidebar():
    """Parse sidebar.html into two-level navigation groups."""
    soup = BeautifulSoup(SIDEBAR_FILE.read_text(encoding="utf-8"), "html.parser")
    nav = []
    current_group = "未分组"
    for elem in soup.find("nav", class_="sidebar").children:
        if elem.name == "div" and "sb-label" in elem.get("class", []):
            current_group = elem.get_text(strip=True)
        elif elem.name == "a" and "anchor-chip" in elem.get("class", []):
            text = elem.get_text(strip=True)
            # separate data badges from the title
            # e.g. "ACC-01 账户总资产卡inline" -> title, mode
            badge_mode = ""
            badge_other = ""
            for badge in elem.find_all("span", class_="data-badge"):
                badge_text = badge.get_text(strip=True)
                if any("mode-" in cls for cls in badge.get("class", [])):
                    badge_mode = badge_text
                else:
                    badge_other = badge_text
            title = text
            if badge_mode:
                title = title.replace(badge_mode, "").strip()
            if badge_other:
                title = title.replace(badge_other, "").strip()
            # extract card id from href
            href = elem.get("href", "")
            card_id = href.lstrip("#") if href else ""
            nav.append({
                "group": current_group,
                "title": title,
                "card_id": card_id,
                "mode": badge_mode,
                "badge_other": badge_other,
            })
    return nav


def extract_trigger(card_id):
    """Extract trigger description from a card HTML file."""
    # Map card id to filename via section-manifest
    manifest_file = BASE_DIR / "assets" / "js" / "section-manifest.js"
    manifest_text = manifest_file.read_text(encoding="utf-8")
    pattern = re.compile(r"\{ id:\s*'" + re.escape(card_id) + r"'\s*,\s*title:\s*'([^']+)'\s*,\s*file:\s*'([^']+)'\s*\}")
    match = pattern.search(manifest_text)
    if not match:
        return "", "", ""
    title, filename = match.group(1), match.group(2)
    file_path = SECTIONS_DIR / filename
    if not file_path.exists():
        return title, filename, ""
    soup = BeautifulSoup(file_path.read_text(encoding="utf-8"), "html.parser")
    section = soup.find("section", id=card_id)
    if not section:
        section = soup.find("section")
    desc_elem = section.find("div", class_="section-desc") if section else None
    if not desc_elem:
        return title, filename, ""
    trigger_text = desc_elem.get_text(separator="", strip=True)
    # Clean up excessive whitespace/newlines
    trigger_text = re.sub(r"\s+", " ", trigger_text).strip()
    return title, filename, trigger_text


def build_rows(nav):
    rows = []
    for item in nav:
        title, filename, trigger = extract_trigger(item["card_id"])
        rows.append({
            "一级目录": item["group"],
            "二级目录/卡片标题": item["title"],
            "卡片编号": item["card_id"].upper() if item["card_id"] else "",
            "dataMode": item["mode"],
            "其他标签": item["badge_other"],
            "触发内容": trigger,
            "文件路径": f"sections/{filename}" if filename else "",
        })
    return rows


def write_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "卡片导航与触发内容"
    headers = ["一级目录", "二级目录/卡片标题", "卡片编号", "dataMode", "其他标签", "触发内容", "文件路径"]
    ws.append(headers)

    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for row_data in rows:
        ws.append([row_data[h] for h in headers])

    # Apply styling and column widths
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 90
    ws.column_dimensions["G"].width = 35

    # Auto row height for trigger content (wrap_text handles display)
    ws.freeze_panes = "A2"
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


if __name__ == "__main__":
    nav = parse_sidebar()
    rows = build_rows(nav)
    output = write_xlsx(rows)
    print(f"Saved: {output}")
    print(f"Total rows: {len(rows)}")
